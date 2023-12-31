from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404, render
from openpyxl import load_workbook

from .models import Category, Product, Quote

# Create your views here.
def index(request):
    category_list = Category.objects.all()
    product_list = Product.objects.all()
    if request.method == "POST":
        print("post request")
        print(request)
    context = {"product_list": product_list, "category_list": category_list}
    return render(request, 'main_page/index.html', context)

def detail(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    products = category.product_set.all()
    context = {'category': category, 'products': products}
    return render(request, "main_page/detail.html", context)

def product_detail(request, product_id):
    #pass another value send over that value in post data
    product = get_object_or_404(Product, pk=product_id)

    if 'my_Cart' in request.session:
        print("It exists")
    else: 
        print("I have to create a my_cart")
        request.session["my_Cart"] = []

    # receiving data from the views
    if request.method == "POST":
        menu=[]
        menu.append(request.session['my_Cart'])

        flattened = []

        for item in menu:
            if isinstance(item, list):
                flattened.extend(item)
                flattened.append(product_id)
            else:
                flattened.append(item)
                flattened.append(product_id)
        # creating a session
        request.session["my_Cart"] = flattened
        print("The new session: ", request.session['my_Cart'])
        # del request.session['cart']

    context = {"product": product}
    return render(request, "main_page/product_detail.html", context)

def import_from_excel(request):
    if request.method == "GET":
        print("hello world")
    
    if request.method == "POST":

        # IMPORTING EXCEL
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        worksheet = wb["Sheet1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

            excel_data.append(row_data)

            # BEGIN EXTRACTING INFO FROM EXCEL INTO DB
            category_var = row_data[3]
            categoryInstance = Category.objects.get(id=category_var)
            name_var = row_data[0]
            SKU_var = row_data[1]
            desc_var = row_data[2]
            vendor_var = row_data[4]
            price_var = row_data[5]
            specialPrice_var = row_data[6]

            # CREATING THIS OBJECT INTO DB
            Product.objects.create(name= name_var, SKU=SKU_var, description=desc_var, category= categoryInstance, vendor=vendor_var, price = price_var, specialPrice=specialPrice_var)

    return render(request, 'main_page/import_form.html')

def view_cart(request):
    cart_list = request.session['my_Cart']
    print(cart_list)
    items_in_cart = {}
    # makes map of quantity
    for i in cart_list:
        if i not in items_in_cart:
            items_in_cart[i] = 0
        items_in_cart[i] +=1
    dict_keys = items_in_cart.keys()
    product_in_cart_list = []
    prices_in_cart =[]

    for item in cart_list:
        product_in_cart = Product.objects.get(id=item)
        product_in_cart_list.append(product_in_cart)
        prices_in_cart.append(product_in_cart.price)

    # this list only contains the items not being repeated
    prod_list =[] 
    for item in dict_keys:
        prod = Product.objects.filter(id=item)[0]
        prod_list.append(prod)

    total_price_of_cart = sum(float(i) for i in prices_in_cart)


    if request.method == "POST":
        user_email = request.POST.get('email')
        print(user_email)
        # user_email = (value of label)
        product_in_cart = request.session['my_Cart']
        price_sum=total_price_of_cart
        # device_cookie=(gah)

        Quote.objects.create(user_email = user_email, products_in_cart=product_in_cart, price_sum=price_sum, device_cookie="12345absde" )


    context = {"prod_list": prod_list, "cart_Items": items_in_cart, "total_price_of_cart": total_price_of_cart }
    return render(request, 'main_page/viewcart.html', context)